from django.db import transaction
from openpyxl import Workbook, load_workbook
from api.services.xlsx_safety import harden_workbook

from api.models import Candidate
from api.services.candidate_validation import (
    candidate_identity_key, clamp_aadhaar_to_last4, identity_key, revalidate_batch_candidates,
)
from api.services.duplicate_check import preload_duplicate_lookup, run_duplicate_check

# Every column is mandatory - a sheet missing one is reported by name (see missing_columns
# below) rather than silently importing blanks for it, and validate_candidate_values() rejects
# any row where one of these is blank.
TEMPLATE_COLUMNS = [
    'Name', 'Email', 'Mobile', 'Aadhaar Last 4 Digits', 'College Name', 'Degree',
    'Stream', 'Percentage', 'Passing Out Year', 'Location',
]

# Nothing is optional at the column level any more - kept as a named set (rather than deleted
# outright) because parse_uploaded_workbook's missing-column check reads it directly, and an
# empty set there reads as "every column is required" just as clearly as removing the check.
OPTIONAL_COLUMNS = set()

# Deliberately generous: recruiters export these sheets from a dozen different systems, and
# a header this map doesn't recognise means the whole column silently imports as blank - which
# looks to the reviewer like every candidate is missing a name rather than like a bad header.
_HEADER_MAP = {
    'name': 'name',
    'full name': 'name',
    'candidate name': 'name',
    'student name': 'name',
    'candidate': 'name',
    'email': 'email',
    'email id': 'email',
    'email address': 'email',
    'e-mail': 'email',
    'mail id': 'email',
    'mobile': 'phone',
    'mobile number': 'phone',
    'mobile no': 'phone',
    'mobile no.': 'phone',
    'phone': 'phone',
    'phone number': 'phone',
    'contact': 'phone',
    'contact number': 'phone',
    'contact no': 'phone',
    'aadhaar last 4 digits': 'aadhaar_last4',
    'aadhaar last 4': 'aadhaar_last4',
    'aadhaar last four digits': 'aadhaar_last4',
    'adhar last 4 digits': 'aadhaar_last4',
    'adhaar last 4 digits': 'aadhaar_last4',
    'aadhaar number': 'aadhaar_last4',
    'aadhaar no': 'aadhaar_last4',
    'aadhaar no.': 'aadhaar_last4',
    'aadhaar': 'aadhaar_last4',
    # "Aadhar" (one 'a') is at least as common in practice as the official spelling.
    'aadhar number': 'aadhaar_last4',
    'aadhar no': 'aadhaar_last4',
    'aadhar no.': 'aadhaar_last4',
    'aadhar': 'aadhaar_last4',
    'adhaar': 'aadhaar_last4',
    'adhar': 'aadhaar_last4',
    'uid': 'aadhaar_last4',
    'college name': 'college_name',
    'college': 'college_name',
    'institute': 'college_name',
    'institution': 'college_name',
    'university': 'college_name',
    'degree': 'degree',
    'qualification': 'degree',
    'stream': 'stream',
    'branch': 'stream',
    'specialization': 'stream',
    'percentage': 'percentage',
    'percent': 'percentage',
    'passing out year': 'passing_out_year',
    'passing year': 'passing_out_year',
    'year of passing': 'passing_out_year',
    'graduation year': 'passing_out_year',
    'batch year': 'passing_out_year',
    'location': 'location',
    'city': 'location',
}

# Which template column each mapped field came from, so a sheet that's missing one can name
# it the way the reviewer sees it in the template.
_FIELD_TO_COLUMN = {
    'name': 'Name',
    'email': 'Email',
    'phone': 'Mobile',
    'aadhaar_last4': 'Aadhaar Last 4 Digits',
    'college_name': 'College Name',
    'degree': 'Degree',
    'stream': 'Stream',
    'percentage': 'Percentage',
    'passing_out_year': 'Passing Out Year',
    'location': 'Location',
}


def generate_template_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Candidates'
    ws.append(TEMPLATE_COLUMNS)
    # The Aadhaar example is 4 digits, matching what the column actually stores (aadhaar_last4
    # is a varchar(4)) - it used to show a full 12-digit number, which uploaded this exact
    # template's own sample row straight into a DataError ("value too long for type character
    # varying(4)"), surfaced to the user as an opaque "could not read that file".
    ws.append(['Jane Doe', 'jane.doe@example.com', '9876543210', '1234',
               'XYZ College', 'B.Tech', 'Computer Science', '78.5', '2025', 'Pune'])
    return harden_workbook(wb)


VALIDATION_REPORT_COLUMNS = [
    'Row', 'Name', 'Email', 'Mobile', 'Aadhaar Last 4 Digits', 'College Name',
    'Status', 'Fields At Fault', 'Errors',
]


def generate_validation_report_workbook(candidates):
    """The invalid rows of an upload, with every error spelled out.

    Only the last 4 Aadhaar digits exist at all now, so there is nothing left to mask - the
    report exists to tell someone which
    rows to fix, and naming the column is enough for that.
    """

    wb = Workbook()
    ws = wb.active
    ws.title = 'Validation Errors'
    ws.append(VALIDATION_REPORT_COLUMNS)
    for candidate in candidates:
        errors = candidate.validation_errors or []
        ws.append([
            candidate.upload_row_number,
            candidate.full_name,
            candidate.email,
            candidate.phone,
            candidate.aadhaar_last4,
            candidate.college_name,
            candidate.get_validation_status_display(),
            ', '.join(dict.fromkeys(e['field'] for e in errors)),
            ' '.join(e['message'] for e in errors),
        ])
    return harden_workbook(wb)


EXPORT_COLUMNS = [
    'Name', 'Email', 'Mobile', 'Batch Name', 'College', 'Degree', 'Stream', 'Percentage',
    'Passing Out Year', 'Location', 'Status', 'Result', 'Overall Score',
]


def generate_candidates_workbook(candidates, latest_attempt_fn, status_display_fn=None):
    """Export for the All Candidates screen. latest_attempt_fn resolves a candidate's
    latest ExamAttempt (or None) - passed in rather than imported, to avoid a serializers
    -> services import cycle. status_display_fn is passed the same way and resolves the
    same derived Status the on-screen table shows (exam attempt, else latest outreach);
    without it the export would fall back to the raw stored status and disagree with the UI.
    """
    if status_display_fn is None:
        status_display_fn = lambda c: c.get_status_display()  # noqa: E731
    wb = Workbook()
    ws = wb.active
    ws.title = 'Candidates'
    ws.append(EXPORT_COLUMNS)
    for candidate in candidates:
        attempt = latest_attempt_fn(candidate)
        overall_score = attempt.overall_score if attempt else candidate.overall_score
        ws.append([
            candidate.full_name,
            candidate.email,
            candidate.phone,
            candidate.batch.batch_name,
            candidate.college_name,
            candidate.degree,
            candidate.stream,
            float(candidate.percentage) if candidate.percentage is not None else None,
            candidate.passing_out_year,
            candidate.location,
            status_display_fn(candidate),
            candidate.get_result_display(),
            float(overall_score) if overall_score is not None else None,
        ])
    return harden_workbook(wb)


def _normalize_header(value):
    if value is None:
        return ''
    # Sheets in the wild label required columns "Email *" or "Email:" - neither changes
    # which column it is.
    return str(value).strip().lower().rstrip('*: ').strip()


def _cell_to_text(value):
    """Excel hands back numbers as floats, so a 4-digit Aadhaar suffix arrives as
    1234.0 and a passing-out year as 2025.0. Rendering those straight through
    str() produced values that failed every format check for no reason the reviewer
    could see - trim the meaningless trailing .0 on whole numbers.
    """
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_uploaded_workbook(file_obj):
    """Parse an uploaded candidate sheet.

    Returns (rows, missing_columns): one dict per data row (1-indexed row_number matching
    the spreadsheet, header = row 1), plus the required template columns whose header this
    sheet didn't carry under any recognised spelling.

    Reporting the missing headers matters more than it looks: an unrecognised "Candidate Name"
    header used to import the whole column as blank, so the review screen showed every single
    row as "Missing Name" and the actual fault - one wrong header cell - was invisible.
    """
    wb = load_workbook(file_obj, data_only=True, read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)

    header_row = next(rows, None)
    if not header_row:
        return [], list(dict.fromkeys(_FIELD_TO_COLUMN.values()))

    field_by_col = {}
    for idx, header in enumerate(header_row):
        field = _HEADER_MAP.get(_normalize_header(header))
        if field and field not in field_by_col.values():
            field_by_col[idx] = field

    found_columns = {_FIELD_TO_COLUMN[field] for field in field_by_col.values()}
    missing_columns = [
        column for column in TEMPLATE_COLUMNS
        if column not in found_columns and column not in OPTIONAL_COLUMNS
    ]

    parsed = []
    for row_number, row in enumerate(rows, start=2):
        if row is None or all(cell in (None, '') for cell in row):
            continue
        data = {'row_number': row_number}
        for idx, field in field_by_col.items():
            data[field] = _cell_to_text(row[idx] if idx < len(row) else None)
        parsed.append(data)
    return parsed, missing_columns


def _to_float(value):
    try:
        return float(value) if value else None
    except ValueError:
        return None


def _to_int(value):
    try:
        return int(float(value)) if value else None
    except ValueError:
        return None


@transaction.atomic
def stage_candidates_from_workbook(batch, file_obj, user, cooling_off_months=3):
    """Parse the uploaded workbook and create one Candidate row per data row, attached
    to `batch` (expected to be in Batch.Status.DRAFT), each validated and duplicate-checked.

    Returns (created, missing_columns, staged, skipped_duplicates).

    Repeated entries collapse to ONE. A sheet listing the same person twice - same Aadhaar, or
    same email - stages only the first occurrence; later copies are skipped and reported rather
    than becoming candidate rows that both get invited and both sit the assessment. Matching is
    seeded from the rows already on the batch too, so adding a second file to the same draft
    dedupes against the first rather than only within itself.

    Only non-empty keys collapse: plenty of rows can legitimately share a blank email, and
    treating "no email" as a match would silently merge unrelated people.

    Wrapped in one transaction so a failure partway through (e.g. a bad row triggering an
    unexpected DB error) can't leave a batch half-populated - the caller's `except Exception`
    around this call sees a clean rollback either way. Rows are parsed into a list up front
    (rather than streamed) so every Aadhaar number is known before duplicate-checking starts,
    letting preload_duplicate_lookup fetch every existing match in one query instead of one
    query per row.

    Validation is deliberately NOT done inline here: it runs once over the finished set via
    revalidate_batch_candidates, because duplicate-email detection needs to see every row
    before it can tell which addresses repeat.
    """
    rows, missing_columns = parse_uploaded_workbook(file_obj)
    duplicate_lookup = preload_duplicate_lookup(
        (row.get('aadhaar_last4') or '').strip() for row in rows
    )

    # Seeded from what's already on the batch so a second upload into the same draft collapses
    # against the first, not just against itself.
    #
    # Keyed on (Aadhaar last 4 + name) rather than the digits alone: with only 4 digits stored,
    # collapsing on the suffix by itself would silently DROP unrelated candidates who happen
    # to share one - a far worse failure than letting a near-duplicate through to review.
    seen_identity, seen_email = set(), set()
    for existing in (batch.candidate_set.filter(is_deleted=False)
                     .only('aadhaar_last4', 'first_name', 'last_name', 'email')):
        if (existing.aadhaar_last4 or '').strip():
            seen_identity.add(candidate_identity_key(existing))
        if existing.email.strip():
            seen_email.add(existing.email.strip().lower())

    created = []
    skipped_duplicates = []
    for data in rows:
        name = data.get('name', '') or ''
        name_parts = name.split(None, 1)
        first_name = name_parts[0] if name_parts else ''
        last_name = name_parts[1] if len(name_parts) > 1 else ''
        # Clamped BEFORE it's used for anything - the identity key below, duplicate matching,
        # and the Candidate.objects.create() call further down all need to see the same,
        # DB-safe value. See clamp_aadhaar_to_last4 for why this exists at all.
        aadhaar = clamp_aadhaar_to_last4((data.get('aadhaar_last4') or '').strip())
        email = (data.get('email') or '').strip()
        row_identity = identity_key({
            'aadhaar_last4': aadhaar, 'first_name': first_name, 'last_name': last_name,
        })

        # Only one row per person. Name + Aadhaar suffix decides first; a repeated email is
        # treated the same way, since the same address twice is the same inbox.
        matched_on = None
        if aadhaar and row_identity in seen_identity:
            matched_on = 'Name + Aadhaar Last 4 Digits'
        elif email and email.lower() in seen_email:
            matched_on = 'Email'
        if matched_on:
            skipped_duplicates.append({
                'row_number': data['row_number'],
                'name': name or '(no name)',
                'email': email,
                'matched_on': matched_on,
            })
            continue
        if aadhaar:
            seen_identity.add(row_identity)
        if email:
            seen_email.add(email.lower())

        candidate = Candidate.objects.create(
            batch=batch,
            upload_row_number=data['row_number'],
            first_name=first_name,
            last_name=last_name or None,
            email=data.get('email', '') or '',
            phone=data.get('phone') or None,
            # The already-clamped local, not a fresh read of data['aadhaar_last4'] - re-reading
            # here would silently undo the clamp above and bring the DataError crash right back.
            aadhaar_last4=aadhaar,
            college_name=data.get('college_name') or None,
            degree=data.get('degree') or None,
            stream=data.get('stream') or None,
            percentage=_to_float(data.get('percentage')),
            passing_out_year=_to_int(data.get('passing_out_year')),
            location=data.get('location') or None,
            # Keep the original text for the two numeric columns. Decimal/SmallInteger can't
            # hold "abc", so it parses to NULL above and would otherwise be indistinguishable
            # from an empty cell - meaning "must be a number" could never be reported.
            upload_raw={
                'percentage': data.get('percentage') or '',
                'passing_out_year': data.get('passing_out_year') or '',
            },
            created_by=user,
        )
        run_duplicate_check(candidate, cooling_off_months=cooling_off_months, existing_lookup=duplicate_lookup)
        if candidate.aadhaar_last4:
            # Keep the lookup current so a later row in this SAME upload that repeats a
            # person is still caught as a duplicate against this one, not just against
            # candidates that existed before the upload started.
            duplicate_lookup.setdefault(candidate_identity_key(candidate), []).insert(0, candidate)
        created.append(candidate)

    # Over the whole batch, not just this upload: a second file added to the same draft can
    # repeat an address from the first.
    staged = revalidate_batch_candidates(batch)
    return created, missing_columns, staged, skipped_duplicates
