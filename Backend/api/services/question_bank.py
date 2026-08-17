from django.db import transaction
from openpyxl import Workbook, load_workbook

from api.models import Question, QuestionBankSection
from api.serializers.question import generate_question_code, normalize_question_text

# The generated template puts each section on its own SHEET, named after the section, so the
# Section column isn't needed - one sheet per section is how these files are actually prepared.
# A Section column is still honoured if a sheet carries one (older files, or a single sheet
# mixing sections), and it overrides the sheet name for that row.
TEMPLATE_COLUMNS = [
    'Question Text', 'Option A', 'Option B', 'Option C', 'Option D',
    'Correct Option', 'Difficulty', 'Marks',
]

_HEADER_MAP = {
    'section': 'section_key',
    'question text': 'question_text',
    'question': 'question_text',
    'option a': 'option_a',
    'option b': 'option_b',
    'option c': 'option_c',
    'option d': 'option_d',
    'correct option': 'correct_option',
    'correct answer': 'correct_option',
    'difficulty': 'difficulty',
    'marks': 'marks',
}

_VALID_DIFFICULTIES = {c[0] for c in Question.Difficulty.choices}
_VALID_OPTIONS = {'A', 'B', 'C', 'D'}


_SAMPLE_ROWS = {
    'logical': [
        'If all Bloops are Razzies and all Razzies are Lazzies, are all Bloops definitely '
        'Lazzies?',
        'Yes', 'No', 'Cannot be determined', 'Only sometimes', 'A', 'Medium', 1,
    ],
    'quantitative': [
        'What is 15% of 240?',
        '36', '32', '40', '24', 'A', 'Easy', 1,
    ],
    'verbal': [
        'Choose the word most similar in meaning to "Candid".',
        'Honest', 'Secretive', 'Aggressive', 'Cautious', 'A', 'Easy', 1,
    ],
    'programming': [
        'Which Python keyword is used to define a function?',
        'def', 'func', 'lambda', 'define', 'A', 'Easy', 1,
    ],
}


def generate_question_template_workbook():
    """One sheet per section, each named after the section.

    Sheet-per-section is how these files are actually prepared, so the template matches: fill
    the Programming sheet with programming questions and they're filed there, with no Section
    column to keep in sync row by row.
    """
    wb = Workbook()
    wb.remove(wb.active)  # replaced by the per-section sheets below

    sections = list(QuestionBankSection.objects.all())
    for section in sections:
        # Excel caps sheet names at 31 characters and forbids : \ / ? * [ ]. A truncated or
        # scrubbed name would no longer match its section on the way back in, so fall back to
        # the short section_key - which the parser accepts as a sheet name too - rather than
        # shipping a template whose own sheets fail validation.
        title = section.section_name
        scrubbed = title
        for bad in ':\\/?*[]':
            scrubbed = scrubbed.replace(bad, ' ')
        if len(scrubbed) > 31 or scrubbed != title:
            title = section.section_key[:31]
        ws = wb.create_sheet(title=title.strip() or section.section_key[:31])
        ws.append(TEMPLATE_COLUMNS)
        sample = _SAMPLE_ROWS.get(section.section_key)
        if sample:
            ws.append(sample)
        ws.freeze_panes = 'A2'

    if not sections:
        # No sections configured yet - still hand back a usable single sheet rather than an
        # empty workbook Excel refuses to open.
        ws = wb.create_sheet(title='Questions')
        ws.append(TEMPLATE_COLUMNS)
        ws.freeze_panes = 'A2'
    return wb


def _normalize_header(value):
    return str(value).strip().lower() if value is not None else ''


def _parse_question_workbook(file_obj):
    """Yields one dict per data row, across EVERY sheet in the workbook.

    Each sheet's name names the section its rows belong to, so a four-section upload is four
    sheets with no Section column to maintain. A row that does carry a Section value overrides
    its sheet, which keeps older single-sheet files working unchanged.

    Row identity: `row_number` is a running index across the whole workbook, because row 2
    exists in every sheet and the validation screen needs one stable id per row to edit and
    remove against. `sheet` and `sheet_row` carry the real location for display.
    """
    wb = load_workbook(file_obj, data_only=True, read_only=True)
    row_number = 0

    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            continue

        field_by_col = {}
        for idx, header in enumerate(header_row):
            field = _HEADER_MAP.get(_normalize_header(header))
            if field:
                field_by_col[idx] = field
        if 'question_text' not in field_by_col.values():
            # Not a question sheet (an instructions tab, a stray blank sheet) - skip it rather
            # than reporting a screenful of "Question Text is required" against it.
            continue

        for sheet_row, row in enumerate(rows, start=2):
            if row is None or all(cell in (None, '') for cell in row):
                continue
            row_number += 1
            data = {
                'row_number': row_number,
                'sheet': ws.title,
                'sheet_row': sheet_row,
            }
            for idx, field in field_by_col.items():
                value = row[idx] if idx < len(row) else None
                data[field] = str(value).strip() if value is not None else ''
            yield data


def _validate_row(data, sections_by_key, valid_section_names, seen_texts):
    """Validate one parsed row. Returns the row dict the Question Validation table renders."""
    errors, error_fields = [], []

    def fail(message, field):
        errors.append(message)
        error_fields.append(field)

    # A row's own Section cell wins; otherwise the sheet it came from names its section. That
    # ordering keeps older single-sheet files (Section column, one sheet) working while making
    # the Section column unnecessary for the sheet-per-section template.
    sheet_name = (data.get('sheet') or '').strip()
    section_name = (data.get('section_key') or '').strip() or sheet_name
    section = sections_by_key.get(section_name.lower())
    if not section_name:
        fail('Section is required.', 'Section')
    elif not section:
        if section_name == sheet_name:
            fail(f'Sheet "{sheet_name}" does not match any section. Rename the sheet to one of: '
                 f'{valid_section_names} - or add a Section column.', 'Section')
        else:
            fail(f'Unknown section "{section_name}". Valid sections: {valid_section_names}.',
                 'Section')

    question_text = (data.get('question_text') or '').strip()
    if not question_text:
        fail('Question Text is required.', 'Question Text')

    option_a = (data.get('option_a') or '').strip()
    option_b = (data.get('option_b') or '').strip()
    option_c = (data.get('option_c') or '').strip()
    option_d = (data.get('option_d') or '').strip()
    if not option_a:
        fail('Option A is required.', 'Option A')
    if not option_b:
        fail('Option B is required.', 'Option B')

    correct_option = (data.get('correct_option') or '').strip().upper()
    if correct_option not in _VALID_OPTIONS:
        fail('Correct Answer must be A, B, C, or D.', 'Correct Answer')
    elif correct_option == 'C' and not option_c:
        fail('Correct Answer is C but Option C is empty.', 'Option C')
    elif correct_option == 'D' and not option_d:
        fail('Correct Answer is D but Option D is empty.', 'Option D')

    difficulty = (data.get('difficulty') or '').strip().title()
    if difficulty not in _VALID_DIFFICULTIES:
        fail(f'Difficulty must be one of {", ".join(sorted(_VALID_DIFFICULTIES))}.', 'Difficulty')

    marks = None
    try:
        marks = int(float(data.get('marks') or 1))
        if marks < 1:
            fail('Marks must be 1 or more.', 'Marks')
    except (TypeError, ValueError):
        fail('Marks must be a number.', 'Marks')

    # A duplicate is reported separately from a hard validation failure: the row is well-formed,
    # it just already exists, which reads differently to an administrator deciding what to fix.
    duplicate_of = None
    normalized = normalize_question_text(question_text)
    if question_text and normalized in seen_texts:
        duplicate_of = seen_texts[normalized]
        fail(f'Duplicate question - already in the bank as {duplicate_of}.', 'Question Text')

    status = 'valid'
    if errors:
        status = 'duplicate' if duplicate_of and len(errors) == 1 else 'invalid'

    return {
        'row_number': data['row_number'],
        'sheet': sheet_name or None,
        'sheet_row': data.get('sheet_row'),
        'section': section_name,
        'section_name': section.section_name if section else None,
        'question_text': question_text,
        'question_type': 'MCQ',
        'option_a': option_a, 'option_b': option_b,
        'option_c': option_c, 'option_d': option_d,
        'correct_option': correct_option,
        'difficulty': difficulty,
        'marks': marks,
        'status': status,
        'errors': errors,
        'error_fields': error_fields,
        'duplicate_of': duplicate_of,
        'question_code': None,
        '_section_obj': section,
        '_normalized': normalized,
    }


def _validation_context():
    """Section lookup + existing-question index, built once per validation run."""
    sections_by_key = {}
    for section_obj in QuestionBankSection.objects.all():
        # Keyed by both the internal section_key ("verbal") and the human-readable section_name
        # ("verbal ability"), since the template and UI show the display name everywhere.
        sections_by_key[section_obj.section_key.lower()] = section_obj
        sections_by_key[section_obj.section_name.lower()] = section_obj
    valid_section_names = ', '.join(sorted({s.section_name for s in sections_by_key.values()}))

    # Existing question texts, normalised, loaded once - a per-row query would be one network
    # round-trip each. Rows accepted during this run are added as we go, so a file that repeats
    # a question inside itself is caught too, not just clashes with what's already in the bank.
    seen_texts = {
        normalize_question_text(text): code
        for text, code in Question.objects.values_list('question_text', 'question_code')
    }
    return sections_by_key, valid_section_names, seen_texts


def _summarize(rows):
    return {
        'total': len(rows),
        'valid': sum(1 for r in rows if r['status'] == 'valid'),
        'invalid': sum(1 for r in rows if r['status'] == 'invalid'),
        'duplicate': sum(1 for r in rows if r['status'] == 'duplicate'),
    }


@transaction.atomic
def validate_question_rows(raw_rows, user=None, dry_run=True):
    """Same validation as validate_question_workbook, but over rows supplied as JSON rather
    than parsed from a file.

    This is what backs per-field editing on the Question Validation screen: the reviewer
    corrects a section name or a correct-answer letter and the edited rows are sent back here
    to be re-checked. Import re-validates through this same path, so a row that was edited into
    an invalid state - or hand-crafted by calling the API directly - still cannot be written.
    """
    sections_by_key, valid_section_names, seen_texts = _validation_context()

    rows = []
    for index, raw in enumerate(raw_rows, start=1):
        data = {
            'row_number': raw.get('row_number') or index,
            # Carried through so a row keeps showing where it came from after being edited,
            # and so a row whose section comes from its sheet still resolves on revalidation.
            'sheet': raw.get('sheet') or '',
            'sheet_row': raw.get('sheet_row'),
            'section_key': raw.get('section') or '',
            'question_text': raw.get('question_text') or '',
            'option_a': raw.get('option_a') or '',
            'option_b': raw.get('option_b') or '',
            'option_c': raw.get('option_c') or '',
            'option_d': raw.get('option_d') or '',
            'correct_option': raw.get('correct_option') or '',
            'difficulty': raw.get('difficulty') or '',
            'marks': raw.get('marks') if raw.get('marks') not in (None, '') else '',
        }
        rows.append(_process_row(data, sections_by_key, valid_section_names, seen_texts,
                                 user, dry_run))
    return rows, _summarize(rows)


def _process_row(data, sections_by_key, valid_section_names, seen_texts, user, dry_run):
    """Validate one row and, unless this is a dry run, create it when it passes."""
    row = _validate_row(data, sections_by_key, valid_section_names, seen_texts)

    if row['status'] == 'valid' and not dry_run:
        question = Question.objects.create(
            question_code=generate_question_code(),
            section=row['_section_obj'],
            question_text=row['question_text'],
            option_a=row['option_a'],
            option_b=row['option_b'],
            option_c=row['option_c'] or None,
            option_d=row['option_d'] or None,
            correct_option=row['correct_option'],
            difficulty=row['difficulty'],
            marks=row['marks'],
            status=Question.Status.ACTIVE,
            created_by=user,
        )
        row['question_code'] = question.question_code

    if row['status'] == 'valid':
        # Registered even on a dry run, so a later row repeating this text is flagged as an
        # in-file duplicate exactly as it would be on the real import.
        seen_texts[row['_normalized']] = row['question_code'] or 'another row in this file'

    row.pop('_section_obj')
    row.pop('_normalized')
    return row


@transaction.atomic
def validate_question_workbook(file_obj, user=None, dry_run=True):
    """Validate every row of an uploaded question workbook, importing the valid ones unless
    this is a dry run.

    Returns (rows, summary). `rows` is one dict per spreadsheet row carrying the parsed values
    plus `status` ('valid' | 'invalid' | 'duplicate'), `errors`, and `error_fields` naming the
    columns at fault - everything the Question Validation table renders. `summary` counts
    total/valid/invalid/duplicate.

    Every sheet in the workbook is read, and each sheet's NAME names the section its rows
    belong to - so a four-section upload is four sheets and needs no Section column. A row that
    does carry a Section value overrides its sheet, which keeps older single-sheet files (one
    sheet, Section column per row) working unchanged. A sheet name matching no section is a
    row-level validation error naming the sheet, never a silently-created wrong record.

    Validation is identical for the dry run and the real import - the import re-reads and
    re-checks the file rather than trusting anything the client sends back, so a tampered
    payload can't slip an invalid row through.
    """
    sections_by_key, valid_section_names, seen_texts = _validation_context()
    rows = [
        _process_row(data, sections_by_key, valid_section_names, seen_texts, user, dry_run)
        for data in _parse_question_workbook(file_obj)
    ]
    return rows, _summarize(rows)
